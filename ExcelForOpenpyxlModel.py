import openpyxl
import pandas

excel_test = openpyxl.load_workbook("E:\工作文档\Onenote\北投beta测试.xlsx")

Excel_sheet = excel_test.worksheets[0]

for row in Excel_sheet.iter_rows():
    for cell in row:
        if cell.value is not None:
            cell.coordinate, cell.value
        else:
            continue

for columns in Excel_sheet.iter_cols():
    for row in columns:
        if row.value is not None:
            row.value
        else:
            continue

for columns in Excel_sheet.columns:
    for row in columns:
        if row.value is not None:
            row.value
        else:
            continue

Excel_sheet.max_column   # 最大列数
Excel_sheet.max_row   # 最大行数
Excel_sheet.cell(1, 1).value   # 获取某个单元格值
Excel_sheet.delete_rows(1)   # 删除行
Excel_sheet.delete_cols(1)   # 删除列
Excel_sheet.merge_cells('A1:B1')   # 合并单元格
Excel_sheet.merge_cells(start_row='1', end_row='3', start_column='1', end_column='3')   # 合并单元格
